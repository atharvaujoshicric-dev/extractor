from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import requests
from bs4 import BeautifulSoup
import re
import io
import json
import time
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import urllib.parse

app = Flask(__name__)
CORS(app)

HEADERS_LIST = [
    {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate',
        'Connection': 'keep-alive',
    },
    {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-GB,en;q=0.5',
        'Connection': 'keep-alive',
    }
]

DESIGNATION_KEYWORDS = {
    'C-Suite': ['ceo', 'cto', 'coo', 'cfo', 'cmo', 'ciso', 'cpo', 'chief executive', 'chief technology', 'chief operating', 'chief financial', 'chief marketing', 'chief information', 'chief product'],
    'VP / SVP': ['vice president', 'vp ', 'svp', 'evp', 'avp', 'senior vice president', 'executive vice president', 'assistant vice president'],
    'Director': ['director', 'head of', 'head -'],
    'Manager': ['manager', 'mgr'],
    'Engineer': ['engineer', 'developer', 'architect', 'programmer', 'sde', 'swe', 'software'],
    'Sales': ['sales', 'account executive', 'account manager', 'business development', 'bd manager', 'bdm'],
    'Marketing': ['marketing', 'growth', 'seo', 'content', 'brand'],
    'HR / People': ['hr ', 'human resource', 'people', 'recruiter', 'talent', 'hiring'],
    'Finance': ['finance', 'accountant', 'analyst', 'financial'],
    'Operations': ['operations', 'ops', 'supply chain', 'logistics', 'procurement'],
    'Founder': ['founder', 'co-founder', 'cofounder', 'owner', 'proprietor'],
    'Intern / Trainee': ['intern', 'trainee', 'apprentice'],
    'Other': []
}

PHONE_PATTERNS = [
    r'\b(?:\+91[-.\s]?)?[6-9]\d{9}\b',
    r'\b(?:\+91[-.\s]?)?\d{5}[-.\s]?\d{5}\b',
    r'\b(?:0\d{2,4}[-.\s]?)?\d{6,8}\b',
    r'\b\+?[1-9]\d{1,14}\b',
]

EMAIL_PATTERN = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'

def get_headers():
    return random.choice(HEADERS_LIST)

def classify_designation(title):
    if not title:
        return 'Other'
    title_lower = title.lower()
    for category, keywords in DESIGNATION_KEYWORDS.items():
        if category == 'Other':
            continue
        for kw in keywords:
            if kw in title_lower:
                return category
    return 'Other'

def extract_phones(text):
    phones = set()
    for pattern in PHONE_PATTERNS:
        matches = re.findall(pattern, text)
        for m in matches:
            cleaned = re.sub(r'[\s\-\.\(\)]', '', m)
            if len(cleaned) >= 10:
                phones.add(cleaned)
    return list(phones)

def extract_emails(text):
    return list(set(re.findall(EMAIL_PATTERN, text)))

def fetch_linkedin_search(company_name):
    """Search LinkedIn via Google for company employees"""
    contacts = []
    queries = [
        f'site:linkedin.com/in "{company_name}" mobile phone contact',
        f'site:linkedin.com/in "{company_name}" director manager engineer',
        f'"{company_name}" employee contact mobile number designation',
    ]
    
    for query in queries[:2]:
        try:
            url = f"https://www.google.com/search?q={urllib.parse.quote(query)}&num=20"
            resp = requests.get(url, headers=get_headers(), timeout=10)
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, 'lxml')
                for result in soup.select('.g'):
                    title_el = result.select_one('h3')
                    snippet_el = result.select_one('.VwiC3b, .s3v9rd, .yXK7lf')
                    if title_el:
                        title_text = title_el.get_text()
                        snippet_text = snippet_el.get_text() if snippet_el else ''
                        full_text = title_text + ' ' + snippet_text
                        
                        name, designation = parse_linkedin_title(title_text, company_name)
                        phones = extract_phones(full_text)
                        emails = extract_emails(full_text)
                        
                        if name:
                            contacts.append({
                                'name': name,
                                'designation': designation,
                                'category': classify_designation(designation),
                                'phone': phones[0] if phones else '',
                                'email': emails[0] if emails else '',
                                'source': 'LinkedIn/Google',
                                'company': company_name
                            })
            time.sleep(random.uniform(1, 2))
        except Exception as e:
            print(f"LinkedIn search error: {e}")
    
    return contacts

def parse_linkedin_title(title, company_name):
    """Parse name and designation from LinkedIn title"""
    name = ''
    designation = ''
    
    # LinkedIn titles usually look like: "John Doe - Senior Engineer at Company | LinkedIn"
    parts = re.split(r'\s*[-|]\s*', title)
    if len(parts) >= 2:
        name = parts[0].strip()
        rest = ' '.join(parts[1:])
        # Remove "LinkedIn" and company name
        rest = re.sub(r'\bLinkedIn\b', '', rest, flags=re.IGNORECASE)
        rest = re.sub(re.escape(company_name), '', rest, flags=re.IGNORECASE)
        rest = re.sub(r'\bat\b', '', rest, flags=re.IGNORECASE)
        designation = rest.strip().strip('-').strip()
        
        # Validate name (2 words, no numbers)
        if not re.match(r'^[A-Za-z]+ [A-Za-z]+', name):
            name = ''
    
    return name, designation

def scrape_company_website(domain):
    """Scrape company website for contact info"""
    contacts = []
    pages_to_check = [
        f"https://{domain}/team",
        f"https://{domain}/about",
        f"https://{domain}/about-us",
        f"https://{domain}/contact",
        f"https://{domain}/people",
        f"https://{domain}/leadership",
        f"https://www.{domain}/team",
        f"https://www.{domain}/about",
        f"https://www.{domain}/contact",
    ]
    
    for url in pages_to_check[:5]:
        try:
            resp = requests.get(url, headers=get_headers(), timeout=8, allow_redirects=True)
            if resp.status_code != 200:
                continue
            
            soup = BeautifulSoup(resp.text, 'lxml')
            text = soup.get_text(separator=' ')
            
            # Extract phones and emails from page
            phones = extract_phones(text)
            emails = extract_emails(text)
            
            # Look for team member cards/sections
            person_selectors = [
                '.team-member', '.member', '.person', '.staff',
                '.employee', '[class*="team"]', '[class*="member"]',
                '[class*="person"]', '[class*="people"]',
                'article', '.card', '.bio'
            ]
            
            found_people = []
            for sel in person_selectors:
                elements = soup.select(sel)
                if elements and len(elements) <= 50:
                    found_people = elements
                    break
            
            if found_people:
                for el in found_people[:20]:
                    el_text = el.get_text(separator=' ')
                    # Look for name patterns (Title Case words)
                    name_match = re.search(r'\b([A-Z][a-z]+ [A-Z][a-z]+(?:\s[A-Z][a-z]+)?)\b', el_text)
                    # Look for designation near name
                    desig_patterns = list(DESIGNATION_KEYWORDS.keys())
                    desig_match = None
                    for cat, kws in DESIGNATION_KEYWORDS.items():
                        for kw in kws:
                            if kw in el_text.lower():
                                desig_match = extract_designation_context(el_text, kw)
                                break
                        if desig_match:
                            break
                    
                    el_phones = extract_phones(el_text)
                    el_emails = extract_emails(el_text)
                    
                    if name_match:
                        contacts.append({
                            'name': name_match.group(1),
                            'designation': desig_match or '',
                            'category': classify_designation(desig_match or ''),
                            'phone': el_phones[0] if el_phones else (phones[0] if phones else ''),
                            'email': el_emails[0] if el_emails else (emails[0] if emails else ''),
                            'source': url,
                            'company': domain
                        })
            else:
                # Fallback: parse full page text for patterns
                lines = [l.strip() for l in text.split('\n') if l.strip()]
                for i, line in enumerate(lines[:100]):
                    name_match = re.match(r'^([A-Z][a-z]+ [A-Z][a-z]+(?:\s[A-Z][a-z]+)?)$', line)
                    if name_match and i + 1 < len(lines):
                        next_line = lines[i+1]
                        designation = next_line if len(next_line) < 60 else ''
                        ph = extract_phones(' '.join(lines[max(0,i-2):i+5]))
                        em = extract_emails(' '.join(lines[max(0,i-2):i+5]))
                        contacts.append({
                            'name': name_match.group(1),
                            'designation': designation,
                            'category': classify_designation(designation),
                            'phone': ph[0] if ph else '',
                            'email': em[0] if em else '',
                            'source': url,
                            'company': domain
                        })
            
            if contacts:
                break
                
        except Exception as e:
            print(f"Error scraping {url}: {e}")
    
    return contacts

def extract_designation_context(text, keyword):
    """Extract designation phrase around keyword"""
    idx = text.lower().find(keyword)
    if idx == -1:
        return keyword
    start = max(0, idx - 10)
    end = min(len(text), idx + 50)
    chunk = text[start:end].strip()
    # Clean up
    chunk = re.sub(r'\s+', ' ', chunk)
    lines = chunk.split('\n')
    for line in lines:
        if keyword in line.lower():
            return line.strip()[:60]
    return chunk[:60]

def search_justdial(company_name, city=''):
    """Scrape JustDial for company contacts"""
    contacts = []
    try:
        query = f"{company_name} {city}".strip()
        url = f"https://www.justdial.com/functions/fetchbusinfo.php?searchstr={urllib.parse.quote(query)}"
        # JustDial is heavily protected, so we use Google to find them
        google_url = f"https://www.google.com/search?q=site:justdial.com+{urllib.parse.quote(query)}+contact"
        resp = requests.get(google_url, headers=get_headers(), timeout=10)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'lxml')
            phones = extract_phones(soup.get_text())
            emails = extract_emails(soup.get_text())
            if phones or emails:
                contacts.append({
                    'name': company_name,
                    'designation': 'Contact',
                    'category': 'Other',
                    'phone': phones[0] if phones else '',
                    'email': emails[0] if emails else '',
                    'source': 'JustDial',
                    'company': company_name
                })
    except Exception as e:
        print(f"JustDial error: {e}")
    return contacts

def search_google_contacts(company_name, city=''):
    """General Google search for company contacts"""
    contacts = []
    queries = [
        f'"{company_name}" {city} director manager contact mobile number',
        f'"{company_name}" {city} CEO CTO founder email phone',
        f'"{company_name}" team leadership contact details',
        f'"{company_name}" {city} employee contact number',
    ]
    
    for query in queries[:3]:
        try:
            url = f"https://www.google.com/search?q={urllib.parse.quote(query)}&num=20"
            resp = requests.get(url, headers=get_headers(), timeout=10)
            if resp.status_code != 200:
                continue
            
            soup = BeautifulSoup(resp.text, 'lxml')
            full_text = soup.get_text()
            phones = extract_phones(full_text)
            emails = extract_emails(full_text)
            
            # Parse search results
            for result in soup.select('.g'):
                title_el = result.select_one('h3')
                snippet_el = result.select_one('.VwiC3b, .yXK7lf, .s')
                if not title_el:
                    continue
                    
                title_text = title_el.get_text()
                snippet_text = snippet_el.get_text() if snippet_el else ''
                combined = title_text + ' ' + snippet_text
                
                r_phones = extract_phones(combined)
                r_emails = extract_emails(combined)
                
                # Try to extract a person name
                name_match = re.search(r'\b([A-Z][a-z]+ [A-Z][a-z]+)\b', combined)
                desig = ''
                for cat, kws in DESIGNATION_KEYWORDS.items():
                    for kw in kws:
                        if kw in combined.lower():
                            desig = extract_designation_context(combined, kw)
                            break
                    if desig:
                        break
                
                if (r_phones or r_emails) and name_match:
                    contacts.append({
                        'name': name_match.group(1),
                        'designation': desig,
                        'category': classify_designation(desig),
                        'phone': r_phones[0] if r_phones else '',
                        'email': r_emails[0] if r_emails else '',
                        'source': 'Google Search',
                        'company': company_name
                    })
            
            time.sleep(random.uniform(1.5, 3))
        except Exception as e:
            print(f"Google search error: {e}")
    
    return contacts

def deduplicate_contacts(contacts):
    """Remove duplicate contacts"""
    seen = set()
    unique = []
    for c in contacts:
        key = (c['name'].lower().strip(), c['phone'], c['email'])
        if key not in seen and (c['name'] or c['phone'] or c['email']):
            seen.add(key)
            unique.append(c)
    return unique

def generate_mock_data(company_name, city=''):
    """Generate realistic sample contacts for demo when scraping fails"""
    import hashlib
    
    seed = int(hashlib.md5(company_name.encode()).hexdigest()[:8], 16)
    random.seed(seed)
    
    first_names = ['Rajesh', 'Priya', 'Amit', 'Sunita', 'Vikram', 'Anita', 'Suresh', 'Kavita', 
                   'Mohit', 'Deepa', 'Ravi', 'Neha', 'Sanjay', 'Pooja', 'Arun', 'Meena',
                   'Kiran', 'Rohit', 'Shweta', 'Manish']
    last_names = ['Sharma', 'Patel', 'Singh', 'Kumar', 'Verma', 'Mehta', 'Joshi', 'Gupta',
                  'Agarwal', 'Reddy', 'Nair', 'Pillai', 'Iyer', 'Rao', 'Shah']
    
    roles = [
        ('Chief Executive Officer', 'C-Suite'),
        ('Chief Technology Officer', 'C-Suite'),
        ('Chief Financial Officer', 'C-Suite'),
        ('VP Sales', 'VP / SVP'),
        ('VP Marketing', 'VP / SVP'),
        ('Director of Operations', 'Director'),
        ('Head of Engineering', 'Director'),
        ('Senior Manager - HR', 'Manager'),
        ('Product Manager', 'Manager'),
        ('Senior Software Engineer', 'Engineer'),
        ('Lead Developer', 'Engineer'),
        ('Sales Manager', 'Sales'),
        ('Business Development Manager', 'Sales'),
        ('Marketing Manager', 'Marketing'),
        ('Finance Manager', 'Finance'),
        ('Founder', 'Founder'),
    ]
    
    contacts = []
    num_contacts = random.randint(8, 15)
    selected_roles = random.sample(roles, min(num_contacts, len(roles)))
    
    for desig, cat in selected_roles:
        first = random.choice(first_names)
        last = random.choice(last_names)
        name = f"{first} {last}"
        
        ph_start = random.choice(['98', '97', '96', '95', '94', '93', '92', '91', '90', '89', '88', '87', '86', '85', '84', '83', '82', '81', '80', '79', '78', '77', '76', '75', '74', '73', '72', '71', '70'])
        phone = f"+91 {ph_start}{''.join([str(random.randint(0,9)) for _ in range(8)])}"
        
        company_domain = company_name.lower().replace(' ', '').replace('.', '')[:12]
        email_name = f"{first.lower()}.{last.lower()}"
        email = f"{email_name}@{company_domain}.com"
        
        contacts.append({
            'name': name,
            'designation': desig,
            'category': cat,
            'phone': phone,
            'email': email,
            'source': 'Demo Data',
            'company': company_name
        })
    
    random.seed()  # Reset seed
    return contacts

@app.route('/api/extract', methods=['POST'])
def extract_contacts():
    data = request.json
    company_name = data.get('company_name', '').strip()
    domain = data.get('domain', '').strip()
    city = data.get('city', '').strip()
    use_demo = data.get('use_demo', False)
    
    if not company_name:
        return jsonify({'error': 'Company name is required'}), 400
    
    contacts = []
    
    if use_demo:
        contacts = generate_mock_data(company_name, city)
    else:
        # Try real scraping
        if domain:
            clean_domain = domain.replace('https://', '').replace('http://', '').replace('www.', '').split('/')[0]
            website_contacts = scrape_company_website(clean_domain)
            contacts.extend(website_contacts)
        
        if len(contacts) < 3:
            google_contacts = search_google_contacts(company_name, city)
            contacts.extend(google_contacts)
        
        if len(contacts) < 3:
            linkedin_contacts = fetch_linkedin_search(company_name)
            contacts.extend(linkedin_contacts)
        
        # If no real data found, fall back to demo
        if len(contacts) == 0:
            contacts = generate_mock_data(company_name, city)
            for c in contacts:
                c['source'] = 'Generated Sample'
    
    contacts = deduplicate_contacts(contacts)
    
    # Get all unique categories
    categories = sorted(set(c['category'] for c in contacts if c['category']))
    
    return jsonify({
        'contacts': contacts,
        'total': len(contacts),
        'categories': categories,
        'company': company_name
    })

@app.route('/api/export', methods=['POST'])
def export_contacts():
    data = request.json
    contacts = data.get('contacts', [])
    company_name = data.get('company_name', 'Company')
    
    if not contacts:
        return jsonify({'error': 'No contacts to export'}), 400
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    
    # Styles
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    
    cat_fills = {
        'C-Suite': 'FF6B6B',
        'VP / SVP': 'FF8E53',
        'Director': 'FFA500',
        'Manager': '4ECDC4',
        'Engineer': '45B7D1',
        'Sales': '96CEB4',
        'Marketing': 'FFEAA7',
        'HR / People': 'DDA0DD',
        'Finance': '98D8C8',
        'Operations': 'F7DC6F',
        'Founder': 'BB8FCE',
        'Intern / Trainee': 'AED6F1',
        'Other': 'D5D8DC',
    }
    
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Title row
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Contact Extraction Report — {company_name}"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1a1a2e')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Blank row
    ws.append([])
    
    # Headers
    headers = ['#', 'Full Name', 'Designation', 'Category', 'Phone Number', 'Email Address', 'Source']
    ws.append(headers)
    header_row = ws.max_row
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[header_row].height = 22
    
    # Data rows
    for i, contact in enumerate(contacts, 1):
        row = [
            i,
            contact.get('name', ''),
            contact.get('designation', ''),
            contact.get('category', 'Other'),
            contact.get('phone', ''),
            contact.get('email', ''),
            contact.get('source', ''),
        ]
        ws.append(row)
        data_row = ws.max_row
        
        cat = contact.get('category', 'Other')
        fill_color = cat_fills.get(cat, 'D5D8DC')
        cat_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        for col in range(1, 8):
            cell = ws.cell(row=data_row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical='center', horizontal='left' if col > 1 else 'center')
            cell.font = Font(name='Arial', size=10)
            if col == 4:
                cell.fill = cat_fill
                cell.font = Font(name='Arial', size=10, bold=True)
        
        ws.row_dimensions[data_row].height = 18
    
    # Column widths
    col_widths = [5, 25, 35, 18, 18, 30, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15
    
    ws2['A1'] = "Category Breakdown"
    ws2['A1'].font = Font(name='Arial', bold=True, size=13, color='1a1a2e')
    ws2.merge_cells('A1:B1')
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 25
    
    ws2.append(['Category', 'Count'])
    for col in range(1, 3):
        cell = ws2.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    category_counts = {}
    for c in contacts:
        cat = c.get('category', 'Other')
        category_counts[cat] = category_counts.get(cat, 0) + 1
    
    for row_i, (cat, count) in enumerate(sorted(category_counts.items()), 3):
        ws2.cell(row=row_i, column=1, value=cat).font = Font(name='Arial', size=10)
        ws2.cell(row=row_i, column=2, value=count).font = Font(name='Arial', size=10, bold=True)
        fill_color = cat_fills.get(cat, 'D5D8DC')
        ws2.cell(row=row_i, column=1).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws2.cell(row=row_i, column=2).alignment = Alignment(horizontal='center')
    
    ws2.append([])
    ws2.append(['Total Contacts', len(contacts)])
    last_row = ws2.max_row
    ws2.cell(row=last_row, column=1).font = Font(name='Arial', bold=True, size=11)
    ws2.cell(row=last_row, column=2).font = Font(name='Arial', bold=True, size=11)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"contacts_{company_name.replace(' ', '_')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/designations', methods=['GET'])
def get_designations():
    return jsonify({'categories': list(DESIGNATION_KEYWORDS.keys())})

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    app.run(debug=True, port=5000)
